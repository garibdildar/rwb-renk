from datetime import date, timedelta
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORDER = ["Red","Blue","Green","Yellow","Peach","White"]
TR = {"Red":"Kirmizi","Blue":"Mavi","Green":"Yesil","Yellow":"Sari","Peach":"Seftali","White":"Beyaz"}
HEX = {"Red":"DC2626","Blue":"2563EB","Green":"16A34A","Yellow":"EAB308","Peach":"FDBA74","White":"F8FAFC"}
FONTHEX = {"Red":"FFFFFF","Blue":"FFFFFF","Green":"FFFFFF","Yellow":"1A1A1A","Peach":"1A1A1A","White":"1A1A1A"}
ANCHOR = date(2026,5,27); ANCHOR_IDX = 5

def period_start(d):
    t=d
    while t.weekday() not in (2,6): t-=timedelta(days=1)
    return t
def next_boundary(d):
    t=d+timedelta(days=1)
    while t.weekday() not in (2,6): t+=timedelta(days=1)
    return t
def period_counter(d):
    epoch=period_start(date(2025,1,1)); ps=period_start(d); t=epoch; c=0
    while t<=ps:
        c+=1; t=next_boundary(t)
    return c
BASE=period_counter(ANCHOR)
def status_for(d):
    nidx=(ANCHOR_IDX+(period_counter(d)-BASE))%6
    out={}
    for i,c in enumerate(ORDER):
        age=(nidx-i)%6
        out[c]= "YENI (tam fiyat)" if age==0 else "Tam fiyat" if age<=2 else "%50 indirim" if age<=4 else "%75 (son sans)"
    return ORDER[nidx], out

thin=Side(style="thin",color="CBD5E1")
border=Border(left=thin,right=thin,top=thin,bottom=thin)
def fill(h): return PatternFill("solid",fgColor=h)
HEADER=fill("1E293B"); HFONT=Font(bold=True,color="FFFFFF")
STAT_FILL={"YENI (tam fiyat)":fill("FACC15"),"Tam fiyat":fill("E2E8F0"),
           "%50 indirim":fill("FB923C"),"%75 (son sans)":fill("EF4444")}
STAT_FONT={"YENI (tam fiyat)":Font(bold=True,color="1A1A1A"),"Tam fiyat":Font(color="334155"),
           "%50 indirim":Font(bold=True,color="1A1A1A"),"%75 (son sans)":Font(bold=True,color="FFFFFF")}

wb=openpyxl.Workbook()

# ---------- Sheet: Veri (data, every date 2026) ----------
wd=wb.active; wd.title="Veri"
cols=["Tarih","Donem baslangici","Yeni renk"]+[TR[c] for c in ORDER]
wd.append(cols)
for j,_ in enumerate(cols,1):
    cell=wd.cell(1,j); cell.fill=HEADER; cell.font=HFONT; cell.alignment=Alignment(horizontal="center")
d=date(2026,1,1); end=date(2026,12,31)
while d<=end:
    newc,out=status_for(d)
    row=[d, period_start(d), TR[newc]]+[out[c] for c in ORDER]
    wd.append(row)
    r=wd.max_row
    wd.cell(r,1).number_format="yyyy-mm-dd"; wd.cell(r,2).number_format="yyyy-mm-dd"
    for k,c in enumerate(ORDER):
        cell=wd.cell(r,4+k); st=out[c]
        cell.fill=STAT_FILL[st]; cell.font=STAT_FONT[st]; cell.alignment=Alignment(horizontal="center")
    d+=timedelta(days=1)
for j in range(1,len(cols)+1):
    wd.column_dimensions[get_column_letter(j)].width=16 if j<=3 else 14
wd.freeze_panes="A2"

# ---------- Sheet: Bugun / Sorgu (lookup) ----------
ws=wb.create_sheet("Bugun - Sorgu",0)
ws.sheet_view.showGridLines=False
ws["B2"]="RWB Thrift — Renk Etiketi Durumu"; ws["B2"].font=Font(bold=True,size=16,color="1E293B")
ws["B3"]="Asagidaki tarihe bir gun yaz; tablo otomatik guncellenir."; ws["B3"].font=Font(color="64748B",size=10)
ws["B5"]="Tarih:"; ws["B5"].font=Font(bold=True)
ws["C5"]="=TODAY()"; ws["C5"].number_format="yyyy-mm-dd"
ws["C5"].fill=fill("FEF9C3"); ws["C5"].font=Font(bold=True,size=12); ws["C5"].border=border
ws["C5"].alignment=Alignment(horizontal="center")
ws["B6"]="Yeni renk (bugun konulan):"; ws["B6"].font=Font(bold=True)
ws["C6"]='=VLOOKUP($C$5,Veri!$A:$L,3,FALSE)'; ws["C6"].font=Font(bold=True,color="B45309")
# table header
ws["B8"]="Renk"; ws["C8"]="Durum"
for col in ("B8","C8"):
    ws[col].fill=HEADER; ws[col].font=HFONT; ws[col].alignment=Alignment(horizontal="center"); ws[col].border=border
for i,c in enumerate(ORDER):
    r=9+i
    cell=ws.cell(r,2,TR[c]); cell.fill=fill(HEX[c]); cell.font=Font(bold=True,color=FONTHEX[c])
    cell.alignment=Alignment(horizontal="center"); cell.border=border
    sc=ws.cell(r,3); sc.value=f'=VLOOKUP($C$5,Veri!$A:$L,{4+i},FALSE)'
    sc.alignment=Alignment(horizontal="center"); sc.border=border
# conditional formatting on status column
from openpyxl.formatting.rule import CellIsRule
rng="C9:C14"
ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"YENI (tam fiyat)"'],fill=STAT_FILL["YENI (tam fiyat)"],font=STAT_FONT["YENI (tam fiyat)"]))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"Tam fiyat"'],fill=STAT_FILL["Tam fiyat"],font=STAT_FONT["Tam fiyat"]))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"%50 indirim"'],fill=STAT_FILL["%50 indirim"],font=STAT_FONT["%50 indirim"]))
ws.conditional_formatting.add(rng, CellIsRule(operator="equal",formula=['"%75 (son sans)"'],fill=STAT_FILL["%75 (son sans)"],font=STAT_FONT["%75 (son sans)"]))
ws.column_dimensions["A"].width=3; ws.column_dimensions["B"].width=26; ws.column_dimensions["C"].width=26

# ---------- Sheet: Takvim (schedule of periods) ----------
wt=wb.create_sheet("Takvim")
wt.sheet_view.showGridLines=False
wt["A1"]="Donem takvimi — her Carsamba & Pazar yeni renk"; wt["A1"].font=Font(bold=True,size=14,color="1E293B")
hdr=["Donem (baslangic - bitis)","Yeni renk"]+[TR[c] for c in ORDER]
wt.append([])  # row2 spacer
wt.append(hdr)
hr=wt.max_row
for j in range(1,len(hdr)+1):
    cell=wt.cell(hr,j); cell.fill=HEADER; cell.font=HFONT; cell.alignment=Alignment(horizontal="center",wrap_text=True); cell.border=border
# list period starts from May 1 to Aug 31 2026
d=period_start(date(2026,5,1))
while d<=date(2026,8,31):
    pe=next_boundary(d)-timedelta(days=1)
    newc,out=status_for(d)
    label=f"{d.strftime('%d %b')} - {pe.strftime('%d %b')}"
    wt.append([label, TR[newc]]+[out[c] for c in ORDER])
    r=wt.max_row
    nc=wt.cell(r,2); nc.fill=fill(HEX[newc]); nc.font=Font(bold=True,color=FONTHEX[newc]); nc.alignment=Alignment(horizontal="center"); nc.border=border
    wt.cell(r,1).border=border
    for k,c in enumerate(ORDER):
        cell=wt.cell(r,3+k); st=out[c]
        cell.fill=STAT_FILL[st]; cell.font=STAT_FONT[st]; cell.alignment=Alignment(horizontal="center"); cell.border=border
    d=next_boundary(d)
wt.column_dimensions["A"].width=22; wt.column_dimensions["B"].width=12
for k in range(len(ORDER)): wt.column_dimensions[get_column_letter(3+k)].width=13
wt.freeze_panes="A4"

wb.save("RWB-Renk-Takvimi.xlsx")
print("saved. periods sample:")
for dd in [date(2026,5,30), date(2026,5,31), date(2026,6,3)]:
    print(dd, status_for(dd))
